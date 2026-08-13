import csv
from pathlib import Path

from _01_config.settings import TEMP_FOLDER
import shutil

import pandas as pd
from openpyxl import load_workbook

from pathlib import Path


import shutil

def get_free_space_gb(
    path
):
    """
    Return free disk space in GB.
    """

    drive = Path(path).anchor

    free_bytes = shutil.disk_usage(
        drive
    ).free

    return round(
        free_bytes /
        1024 /
        1024 /
        1024,
        2
    )

# def save_to_csv(df, output_file):

#     output_file = Path(output_file)

#     temp_file = output_file.with_suffix(
#         output_file.suffix + ".tmp"
#     )

#     # Write temp file first
#     df.to_csv(
#         temp_file,
#         index=False,
#         encoding="utf-8-sig",
#         quoting=csv.QUOTE_ALL,
#         quotechar='"',
#         doublequote=True
#     )

#     # Replace existing file atomically
#     temp_file.replace(output_file)

#     print(f"Saved: {output_file}")


def save_to_csv(
    df,
    path,
    **kwargs
):
    output_file = Path(path)

    temp_file = output_file.with_suffix(
        output_file.suffix + ".tmp"
    )

    output_file.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    try:
        df.to_csv(
            temp_file,
            index=False,
            encoding="utf-8-sig",
            quoting=csv.QUOTE_ALL,
            quotechar='"',
            doublequote=True,
            **kwargs
        )

        temp_file.replace(output_file)

        print(f"\nSaved: {output_file}")

    except OSError as e:

        if "No space left" in str(e):

            free_space = get_free_space_gb(
                output_file
            )

            raise OSError(
                f"""
Disk Full

Cannot save:

{output_file}

Free Space Remaining:
{free_space} GB
"""
            )

        raise




def save_to_xlsx(df, output_file, sheet_name):

    if not isinstance(df, pd.DataFrame):
        raise TypeError(
            f"Expected pandas.DataFrame, got {type(df).__name__}"
        )

    output_file = Path(output_file)
    output_file.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    # File doesn't exist
    if not output_file.exists():

        with pd.ExcelWriter(
            output_file,
            engine="openpyxl"
        ) as writer:

            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

        return

    # File exists
    wb = load_workbook(output_file)

    try:
        sheet_exists = (
            sheet_name in wb.sheetnames
        )
    finally:
        wb.close()

    with pd.ExcelWriter(
        output_file,
        engine="openpyxl",
        mode="a",
        if_sheet_exists=(
            "replace"
            if sheet_exists
            else "new"
        )
    ) as writer:

        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False
        )