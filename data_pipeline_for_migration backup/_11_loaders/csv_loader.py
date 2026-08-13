from pathlib import Path
from _01_config.settings import TEMP_FOLDER
import shutil

import pandas as pd
from openpyxl import load_workbook




def load_csv(path, low_memory=False, **kwargs):

    return pd.read_csv(
        path,
        low_memory=low_memory,
        quotechar='"',
        **kwargs
    )




def load_xls(
    path,
    sheet_name='sheet 1',
    dtype=None,
    na_filter=True,
    keep_default_na=True,
    parse_dates=False,
    engine="openpyxl",
    data_only=False,
    **kwargs
):
    """
    Safely load an Excel worksheet.

    Parameters
    ----------
    path : str | Path
    sheet_name : str | int
        Worksheet name or index.
    dtype : dict | str | None
    parse_dates : bool | list
    data_only : bool
        If True, returns formula results instead of formulas.
    kwargs :
        Additional arguments passed to pd.read_excel()
    """

    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {path}")

    try:
        if data_only:
            workbook = load_workbook(path, data_only=True, read_only=True)

            if isinstance(sheet_name, str):
                if sheet_name not in workbook.sheetnames:
                    raise ValueError(
                        f"Sheet '{sheet_name}' not found. "
                        f"Available sheets: {workbook.sheetnames}"
                    )

        return pd.read_excel(
            path,
            sheet_name=sheet_name,
            dtype=dtype,
            na_filter=na_filter,
            keep_default_na=keep_default_na,
            parse_dates=parse_dates,
            engine=engine,
            **kwargs
        )

    except Exception as e:
        raise RuntimeError(
            f"Failed to load Excel file '{path}': {e}"
        ) from e
